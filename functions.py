### imports
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime, timedelta
import win32com
import win32com.client
from tkinter import filedialog as fd
import json
import requests
import openpyxl
from openpyxl import load_workbook
import csv
import openpyxl
import pandas as pd
import datetime
import numpy as np
from openpyxl.styles import PatternFill, Font
from win32com.client import Dispatch
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import date


# paths needed for dynamic file storage and location
import os
downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads\&')
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop\&')
bot_path = os.path.join(os.path.expanduser('~'), 'Desktop\&bill_bot-main')

### Import the required libraries
from tkinter import *
from functions import *

### set frames and base grid
# create instance
root = Tk()
# set base dimensions and features
root.geometry("470x500")
root.title('Bill_Bot v1')

### set dimensions for frames
task_frame = LabelFrame(root, text="Tasks", padx=2, pady=2)
task_frame.grid(rowspan=8, row=0, column=0)

menu_frame = LabelFrame(root, padx=5, pady=5)
menu_frame.grid(rowspan=1, row=1, column=1)

### title
label = Label(root, text="Welcome!")
label.grid(row=0, column=1)

### Create dropdown Menus
service_list = ['Posted/Unposted','Focus File','Overdue Invoices','All Daily']
    #             ,\
    # 'Contract Mapping','AR Report','PIE Import','Budget V Spend', 'SST-SBLIW',\
    # 'Quick Pulse','MI45 Reminder','FIWLR','FIWLR w/Miscodes','CP Actuals','EOM AR Report','Cadence Files','OEM Files','Odd Day','Billing Audit','Cummulative Overdue','Planner Checks']
listbox_daily = Listbox(task_frame, width=40, height=24, selectmode=MULTIPLE)

# Inserting the listbox items
for item in service_list:
    listbox_daily.insert(END, item)
listbox_daily.pack()

### function to select files // tied to "Prepare Reports" button
def select_docs():
    # set file names as global so other classes and their objects can reference and edit 
    global selected_files, billing_register, csp_transactions, unposted_invoices, gbs_export, pie_extract, selected_files
    filetypes = (('text files', '*.txt'),('All files', '*.*'))
    selected_files = list(fd.askopenfilenames(
        title='Open files',
        initialdir='/',
        filetypes=filetypes))
    try:
        ### read files selected by user and assign them to variables for program usage
        for i in selected_files:
            i = i.replace("/", "\\")
            if "Extract" in i:
                pie_extract = pd.read_excel(i, skiprows = range(1, 8), usecols = "A:BK")
                pie_extract.rename(columns = pd.Series(pie_extract.iloc[0]), inplace=True)
                pie_extract = pie_extract.iloc[1:-1]

            elif "Register" in i:
                billing_register = pd.read_excel(i, skiprows = range(1, 8))
                billing_register.rename(columns = pd.Series(billing_register.iloc[0]), inplace=True)
                billing_register = billing_register.iloc[1:-1]

            elif "CSP" in i:
                csp_transactions = pd.read_excel(i, skiprows = range(1, 8))
                csp_transactions.rename(columns = pd.Series(csp_transactions.iloc[0]), inplace=True)
                csp_transactions = csp_transactions.dropna().iloc[1:-1]

            elif "Unposted" in i:
                unposted_invoices = pd.read_excel(i, skiprows = range(1, 8))
                unposted_invoices.rename(columns = pd.Series(unposted_invoices.iloc[0]), inplace=True)
                unposted_invoices = unposted_invoices.iloc[1:-1]

            elif "export" in i.lower():
                gbs_export = pd.read_excel(i)
    except:
        print("no files selected yet")
 
### read distribution lists file and clear missing values
all_distributions = pd.read_excel("workingFiles/distributions/all_distributions.xlsx").dropna()
# set distribution lists
test_dir, posted_unposted_dir, focus_file_dir, overdue_invoices_dir = ['rcelisduran@ibm.com'],\
    all_distributions.posted_unposted.values.tolist(),\
    all_distributions.focus_file.values.tolist(),\
    all_distributions.overdue_invoices.values.tolist()

### time and dates for emails and reports
last_friday, yesterday =\
    str(date.today() - timedelta(3)),\
    str(date.today() - timedelta(1))
# if today is Monday the date for the subject will be Friday's date, otherwise it will be yesterday's
if date.today().weekday() != 0:
    subject_date = yesterday
else:
    subject_date = last_friday

# function to download files for less repetition
def download_doc(data, file_name):
    try:
        downloads_path.replace("\&", "\\")
        file_path = downloads_path + file_name + "_" + str(subject_date)
        print(file_path)
        data.to_excel(file_path)
    except:
        print(f"{file_name} - data to excel - FAIL")

### tasks class with sub classes and their functions
class Tasks():
    class t_daily():
        global billing_register, csp_transactions, unposted_invoices, pie_extract

        def posted_unposted():
            df = csp_transactions.merge(billing_register, how='right', on='Invoice ID').merge(pie_extract, on='Project ID').merge(unposted_invoices, how="left", on='Project ID',suffixes=("","_delme"))
            df = df[[c for c in df.columns if not c.endswith('_delme')]]

            print(df.columns)

            new_df = pd.DataFrame()
            new_df['Project ID'], new_df['Fiscal Year'],new_df['Period'], new_df['Invoice ID'], new_df['Invoice Date'], new_df['Invoice Amount'], new_df['CSP']  = \
            df['Project ID'].astype(str),  df['Fiscal Year'], df['Period_y'] ,df['Invoice ID'].astype(str), pd.to_datetime(df['Invoice Date_y'], format=("%m-%d-%Y")), df['Invoice Amount'], df['Total CSP Amount']

            new_df['Net Div 16 Invoice'], new_df['Posted Date'] =\
            new_df['Invoice Amount'] - new_df['CSP'], pd.to_datetime(df['Invoice Posted Date'], format=("%m-%d-%Y"))

            i = 0
            while i < len(list(new_df['Posted Date'])):
                new_df['Posted Date'][i], new_df['Invoice Date'][i] = new_df['Posted Date'][i].date(), new_df['Invoice Date'][i].date()
                i += 1

            i = 0
            new_df['Due in GBS'] = new_df['Posted Date']
            while i < len(list(new_df['Posted Date'])):
                # missing due in gbs
                if (new_df['Invoice Date'].iloc[i].weekday() == 4):
                    new_df['Due in GBS'].iloc[i] += datetime.timedelta(4)

                elif (new_df['Invoice Date'].iloc[i].weekday() == 5):
                    new_df['Due in GBS'].iloc[i] += datetime.timedelta(4)

                elif (new_df['Invoice Date'].iloc[i].weekday() == 6):
                    new_df['Due in GBS'].iloc[i] += datetime.timedelta(3)
                    
                else:
                    new_df['Due in GBS'].iloc[i] += datetime.timedelta(2)
                i += 1

            new_df['In GBS?'] = None
            i = 0
            # "In GBS?" column made from unposted report
            while i < len(list(new_df['Invoice ID'])):
                if new_df['Invoice ID'][i] in (str(list(gbs_export['Invoice Number']))):
                    new_df['In GBS?'][i] = "Yes"
                else:
                    if new_df['Invoice ID'][i] in (str(list(gbs_export['CCI-Invoice-Numbers']))):
                        new_df['In GBS?'][i] = "Yes"
                    else:
                        new_df['In GBS?'][i] = "No"   
                i += 1

            # fa col
            new_df['Invoicer'] = df['Financial Analyst_y']

            # work days past posted
            new_df['Work Days Past Posted Date'] = None
            i = 0
            while i < len(list(new_df['In GBS?'])):
                if new_df['In GBS?'][i] == "No":
                    new_df['Work Days Past Posted Date'][i] = np.busday_count(new_df['Due in GBS'][i], datetime.datetime.now().date())
                    if new_df['Work Days Past Posted Date'][i] <= 0:
                        new_df['Work Days Past Posted Date'][i] = None
                else:
                    new_df['Work Days Past Posted Date'][i] = "Submitted in repository"
                i+=1

            for i in range(0,len(new_df['Project ID'])):
                if "DDOU." in str(new_df['Project ID'][i]):
                    new_df = new_df.drop(i)
                else:
                    print('check')
           
            base_path = downloads_path.replace("\&", "\\")

            print(base_path)
            base_path += "\\posted_unposted_" + str(subject_date) + '.xlsx'
            print(base_path)

            import openpyxl
            base_excel = new_df.to_excel(base_path, index=False)

            workbook = openpyxl.load_workbook(base_path)
            wb = workbook.active

            font = Font(bold=True)
            # highlight header
            fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            for column in wb.iter_cols():
                if column[0].value is not None:
                    cell = column[0]
                    cell.font = font
                    cell.fill = fill

            # define number format
            dollar_format = '$###,###,##0.00'
            for column in ['F', 'G', 'H']:
                for cell in wb[column]:
                    cell.number_format = dollar_format

            # align cells
            for cell in wb[1]:
                if cell.value is not None:
                    column_letter = cell.column_letter
                    for cell in wb[column_letter]:
                        cell.alignment = Alignment(horizontal='center')

            try:
                fill_overdue = PatternFill(start_color='FF7D40', end_color='FF7D40', fill_type='solid')
                count = 1
                for col in wb.iter_cols():
                    if col == 'M':
                        for cell in wb.iter_rows():
                            if cell.value > 0:
                                for col in wb.iter_cols():
                                    col[count].font = font
                                    col[count].fill = fill_overdue
                                count += 1
                            else: 
                                count += 1
            except:
                print('highlighting overdues non-functional')
            
            tracking = pd.DataFrame(),
            focus_extract = pd.DataFrame()
            overdue_file = pd.DataFrame()
            gbs_import = pd.DataFrame()

            try:
                gbs_import = new_df[['Project ID', 'Invoice ID', 'Invoice Date', 'Invoice Amount', 'In GBS?']]
                download_doc(data=gbs_import, file_name='gbs_import')
            except:
                print('gbs import creation - FAILED')

            try:
                focus_extract = df[['Division', 'Fiscal Year', 'Period', 'Invoice ID']]
                focus_extract['Invoice Date'], focus_extract['Invoice Posted'], focus_extract['Due in GBS'], focus_extract['In GBS?'] \
                    = new_df['Invoice Date'], new_df['Invoice Posted'], new_df['Due in GBS'], new_df['In GBS?']
                focus_extract['CMR'], focus_extract['Project ID'], focus_extract['Billing Format'], focus_extract['Project Name']\
                      = df['CMR'], df['Project ID'], df['Billing Format'], df['Project Name']
                focus_extract['Customer Name'], focus_extract['Key Account'], focus_extract['Sub Account'] \
                    = df['Customer Name'], df['Key Account'], df['Sub Account']
                focus_extract['Porject Manager'], focus_extract['Financial Analyst'], focus_extract['Billing Grp'], focus_extract['Invoicer'] \
                    = df['Project Manager'], df['Financial Analyst'], df['Billing Grp'], df['Invoicer']
                focus_extract['Invoice Amount'], focus_extract['CSP'], focus_extract['Net Invoice'], focus_extract['Transmitted to CCS'] \
                    = df['Invoice Amount'], df['CSP'], df['Invoice Amount'] - df['CSP'], df['Transmitted to CCS']

                download_doc(data=focus_extract, file_name='focus_extract')
            except:
                print('focus creation - FAILED')

            try:
                overdue_file = new_df[['Project ID', 'Invoice ID', 'Net Div 16 Invoice', 'Posted Date', 'Due in GBS', 'In GBS?', 'Invoicer', 'Work Days Past Posted Date']]
                overdue_file = overdue_file[overdue_file['Work Days Past Posted Date'] > 0]

                download_doc(data=overdue_file, file_name='overdue_invoices')
            except:
                print('overdue creation - FAILED')

            # save changes 
            workbook.save(base_path)
            workbook.close()

            excel = Dispatch('Excel.Application')
            wb = excel.Workbooks.Open(base_path)
            excel.Worksheets(1).Activate()
            excel.ActiveSheet.Columns.AutoFit()
            # save and close file
            wb.Save()
            wb.Close()

            outlook = win32com.client.Dispatch("Outlook.Application")
            message = outlook.CreateItem(0)

            message.To = "; ".join(test_dir)
            message.Subject = ("Posted Invoices as of 7:00 PM EST "+subject_date)
            message.Body = "Attached are the posted/unposted invoices as of 7pm EST " + subject_date
            message.Attachments.Add(Source = base_path)
            message.Send()

### function to activate task functions // tied to "Prepare Reports" button
def perform_tasks(task_list):
    # perform tasks (call functions going down list of tasks)
    for task in task_list:
        if task == "Posted/Unposted":
            print(task_list, selected_files)
            Tasks.t_daily.posted_unposted()
        elif task == "Focus File":
            Tasks.t_daily.focus_file()
        elif task ==  "Overdue Invoices":
            Tasks.t_daily.overdue_invoices()
        elif task == "All Daily":
            Tasks.t_daily()

### Function to process selected requests
tasks_list = []
def submit_requests():
    label = Label(root, text="")
    for i in listbox_daily.curselection():
        tasks_list.append(listbox_daily.get(i))
    if tasks_list == []:
        label = Label(root, text="Please select your tasks and try again...")
        label.grid(row=0, column=1)
    else:
        select_docs()
        label = Label(root, text="Your requests are being processed...")
        label.grid(row=0, column=1)
        perform_tasks(task_list=tasks_list)

# button to process reports
submit_button = Button(\
    menu_frame,\
    text="Prepare Reports",\
    padx=50,\
    command=submit_requests)\
    .grid(row=4,column=10)

### needed for app
root.mainloop()
